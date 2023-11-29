import openpyxl
import requests
import re
import xlwt
from bs4 import BeautifulSoup

hd = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36'
}


def getmanget(linkurl):
    res = requests.get(linkurl, headers=hd)
    res.encoding = 'GBK'
    soup = BeautifulSoup(res.text, 'html.parser')
    ret = soup.find_all('a')
    for n in ret:
        if 'magnet' in str(n.string):
            return n.string
    return None

def saveExcel( data,page):
    # 创建一个新的Excel文档
    workbook = openpyxl.Workbook()

    # 选择默认的工作表
    sheet = workbook.active
    for i in range(len(data)):
        sheet.cell(row=i+1, column=1, value=data[i])
    workbook.save('./guochan/movie'+str(page)+'.xlsx')
    workbook.close()

count = 0
workbook = xlwt.Workbook(encoding='utf-8')
worksheet = workbook.add_sheet('sheet1')

for i in range(1,2):
    if i == 1:
        url = 'https://www.dygod.net/html/gndy/china/index.html'
    else:
        url = 'https://www.dygod.net/html/gndy/china/index_' + str(i) + '.html'
    # url = 'https://www.dygod.net/html/tv/rihantv/index_' + str(i) + '.html'
    res = requests.get(url, headers=hd,verify=False)
    res.encoding = 'GBK'
    soup = BeautifulSoup(res.text, 'html.parser')
    ret = soup.find_all(class_='tbspan', style='margin-top:6px')
    info = []
    for x in ret:

        a_tags = x.find_all("a")

        info.append(a_tags[1].string)

        linkurl = 'https://www.dygod.net/' + a_tags[1].get("href")
        manget = getmanget(linkurl)
        print(manget)
        if manget:
            info.append(str(manget).split('&')[0])
    saveExcel(info,i)