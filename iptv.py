import openpyxl
import requests
import re
import xlwt
from bs4 import BeautifulSoup

hd = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/30.0.1599.101 Safari/537.36'
}


def getmanget(linkurl,info):
    res = requests.get(linkurl, headers=hd)
    res.encoding = 'GBK'
    soup = BeautifulSoup(res.text, 'html.parser')
    ret = soup.find_all('a')
    for n in ret:
        if 'magnet' in str(n.string):
            info.append(str(n.string).split('&')[0])

def saveExcel( data,page):
    # 创建一个新的Excel文档
    workbook = openpyxl.Workbook()

    # 选择默认的工作表
    sheet = workbook.active
    for i in range(len(data)):
        sheet.cell(row=i+1, column=1, value=data[i])
    workbook.save('tvb'+str(page)+'.xlsx')
    workbook.close()

count = 0
workbook = xlwt.Workbook(encoding='utf-8')
worksheet = workbook.add_sheet('sheet1')

for i in range(2,6):
    url = 'https://www.dygod.net/html/dongman/index_' + str(i) + '.html'
    # url = 'https://www.dygod.net/html/tv/rihantv/index_' + str(i) + '.html'
    res = requests.get(url, headers=hd,verify=False)
    res.encoding = 'GBK'
    soup = BeautifulSoup(res.text, 'html.parser')
    ret = soup.find_all(class_='tbspan', style='margin-top:6px')
    for x in ret:

        a_tags = x.find_all("a")

        # info.append(a_tags[1].string)

        linkurl = 'https://www.dygod.net/' + a_tags[0].get("href")
        tit = a_tags[0].get("title")
        if "》全" in tit :
            print(tit)
            info = []
            getmanget(linkurl,info)
            print(info)
            saveExcel(info,tit)